import os

from tkinter import *
from tkinter import messagebox

from datetime import datetime

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title("Attendence App")
root.iconphoto(True,PhotoImage(file=r"img\check.png"))
root.geometry("1000x300")

DEFAULT_FONT = font=("Calibri",15)
SECONDARY_FONT = font=("Calibri",18)

EXCEL_FILE_NAME = "student\\student.xlsx"

# Functions
def add_student():
    def submit():
        std = entry.get()
        if not std=="":
            ws.append([std])

            # Save File
            wb.save(EXCEL_FILE_NAME)

            # Message Box
            messagebox.showinfo(title="Attendence App",message="Student Added!")

            # Quit top
            top.destroy()

    # Create Work Book instance
    wb = Workbook()

    # Load existing Workbook
    wb = load_workbook(EXCEL_FILE_NAME)

    # Create Active Worksheet
    ws = wb.active

    # Create New Window
    top = Toplevel(root)
    top.geometry("410x100")

    # Add Label in New Window
    Label(top,text="Enter New Student: ",font=DEFAULT_FONT,padx=10).grid(row=0,column=0)

    # Add Entry in New Window
    entry = Entry(top,font=DEFAULT_FONT)
    entry.grid(row=0,column=1,pady=10)

    # Add Submit Button in New Window
    submit_btn = Button(top,text="Add",command=submit)
    submit_btn.grid(row=1,column=1)

    # Save File
    wb.save(EXCEL_FILE_NAME)

def view_student():
    # Create New Window
    top = Toplevel(root)
    top.geometry("400x300")

    # Define Scrollbar and Textbox
    scrollbar = Scrollbar(top)
    scrollbar.pack(side=RIGHT, fill=Y)
    textbox = Text(top,font=DEFAULT_FONT)
    textbox.pack()
    
    # Create Work Book instance
    wb = Workbook()

    # Load existing Workbook
    wb = load_workbook(EXCEL_FILE_NAME)

    # Create Active Worksheet
    ws = wb.active

    # Create Variable for Column A
    column_a = ws['A']

    # counter var
    i = 1

    for cell in column_a:
        textbox.insert(END,f"{i}. {cell.value}\n")
        i+=1

    # Add Textbox to Scrollbar
    textbox.config(yscrollcommand=scrollbar.set)
    scrollbar.config(command=textbox.yview)

    # Disable Text Box
    textbox["state"] = DISABLED

def get_std_list():
    # Create Work Book instance
    wb = Workbook()

    # Load existing Workbook
    wb = load_workbook(EXCEL_FILE_NAME)

    # Create Active Worksheet
    ws = wb.active

    # Create Variable for Column A
    column_a = ws['A']

    # counter var
    i = 1

    # list
    std_list = []

    for cell in column_a:
        std_list.append(cell.value)
        i+=1
    
    std_list.pop(0)
    return std_list

def attendence_student():
    l = os.listdir("data//")
    cur_month = datetime.now().strftime("%B")

    cur_month_name = f"data//{cur_month}.xlsx"
    def mark_attendence():
        cur_month_name = f"data//{cur_month}.xlsx"
        # Create Work Book instance
        wb = Workbook()

        # Load existing Workbook
        wb = load_workbook(cur_month_name)

        # Set Workbook's Active Sheet
        for i, s in enumerate(wb.sheetnames):
            if s == str(datetime.now().strftime("%d")):
                break
        wb.active = i

        # Set Workbook's Active Sheet to ws
        ws = wb.active

        ws["A1"] = "Names"
        ws["B1"] = "Attendence"
        for i in range(len(std_list)):
            a = str(std_list_vars[i].get())
            name = std_list[i]

            ws[f"A{i+2}"] = name
            ws[f"B{i+2}"] = a

        # Save Workbook
        wb.save(cur_month_name)

        # Message Box
        messagebox.showinfo(title="Attendence App",message="Added Attendence!")

        # Quit top
        top.destroy()

    # Create a Excel book if not there!
    if not f"{cur_month}.xlsx" in l:
        book = Workbook()
        book.save(cur_month_name)

    # Create Work Book instance
    wb = Workbook()

    # Load existing Workbook
    wb = load_workbook(cur_month_name)

    # Create Today's Excel Sheet
    wb.create_sheet(datetime.now().strftime("%d")) if not datetime.now().strftime("%d") in wb.sheetnames else None

    # Set Workbook's Active Sheet
    for i, s in enumerate(wb.sheetnames):
        if s == str(datetime.now().strftime("%d")):
            break
        wb.active = i

    # Set Workbook's Active Sheet to ws
    ws = wb.active

    # New Window Code
    top = Toplevel(root)
    top.geometry("400x500")

    # Student's List
    std_list = get_std_list()
    print(std_list)

    # Student's List Label & Checkbox & Variable
    std_list_label = []
    std_list_vars = []
    std_list_checkbox = []

    # Add Everything
    for i in range(len(std_list)):
        std_list_label.append(Label(top,text=f"{i+1}. {std_list[i]}",font=DEFAULT_FONT))
        std_list_vars.append(BooleanVar())
        std_list_checkbox.append(Checkbutton(top,variable=std_list_vars[i],font=DEFAULT_FONT))
        std_list_label[i].grid(row=i,column=0,padx=10)
        std_list_checkbox[i].grid(row=i,column=1)
    
    # Add Submit Button
    submit_btn = Button(top,text="Mark!",font=DEFAULT_FONT,command=mark_attendence)
    submit_btn.grid(row=i+1,column=1)
    
    # Delete "Sheet" sheet from wb
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save Workbook
    wb.save(cur_month_name)

# Create Header
Label(root,text="Attendence App",font=("Pacifico",25)).pack()

# Load Images
add_student_img = PhotoImage(file=r"img\add.png")
view_student_img = PhotoImage(file=r"img\view.png")
attendence_student_img = PhotoImage(file=r"img\check.png")

# Image Buttons Frame
image_btn_frame = Frame(root,padx=30,pady=20)
image_btn_frame.pack(fill=BOTH,expand=1)

# Add Button
add_btn = Button(image_btn_frame,image=add_student_img,command=add_student)
add_btn.grid(row=0,column=0,padx=80)

# Add Label
Label(image_btn_frame,text="Add Student",font=SECONDARY_FONT).grid(row=1,column=0,pady=20)

# View Button
view_btn = Button(image_btn_frame,image=view_student_img,command=view_student)
view_btn.grid(row=0,column=1,padx=80)

# View Label
Label(image_btn_frame,text="View Students",font=SECONDARY_FONT).grid(row=1,column=1,pady=20)

# Attendence Button
attendence_btn = Button(image_btn_frame,image=attendence_student_img,command=attendence_student)
attendence_btn.grid(row=0,column=2,padx=80)

# Attendence Label
Label(image_btn_frame,text="Attendence",font=SECONDARY_FONT).grid(row=1,column=2,pady=20)

root.mainloop()